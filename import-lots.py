import openpyxl, datetime

SRC = '/Users/fahdnaveed/Downloads/LOT Analysis.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['LOT Analysis']

# canonical buyer map (uppercased/stripped key -> canonical)
BUYER = {
    'MY FOOD':'My Food','MY FOODS':'My Food',
    'VIMA':'Vima','CORPORACION ALIMENTARIA VIMA S.L.':'Vima',
    'SEA BLUE, UK':'Sea Blue, UK','SEA BLUE,UK':'Sea Blue, UK','SEABLUE, UK':'Sea Blue, UK',
    'SEA PRIDE':'Sea Pride','AL HAMOOR':'Al Hamoor','AH-CHOU':'Ah-Chou',
    'QUIRCH':'Quirch','QUICH':'Quirch','GENERAL':'General','GLOBAL':'Global',
    'OCEANIS':'Oceanis','LUCKY FISH':'Lucky Fish','LUCKY':'Lucky Fish',
    'VASSILIOU':'Vassiliou','SCHROEDER':'Schroeder','THALASA':'Thalassa',
    'THALASSA SEAFOODS NV/SA':'Thalassa','PESCANOVA':'Pescanova','FEINFROST':'Feinfrost',
    'NORDIC':'Nordic','IST':'IST','SEAFOOD':'Seafood','AZHARI':'Azhari',
    'EURL':'EURL','BML':'BML',
}
def clean_buyer(v):
    if v is None: return None, None
    raw = str(v).strip()
    if not raw: return None, None
    key = ' '.join(raw.upper().split())
    return BUYER.get(key, raw.title()), raw

def clean_product(v):
    if v is None: return None
    return ' '.join(str(v).strip().upper().split()) or None

def n(v):
    return float(v) if isinstance(v,(int,float)) and not isinstance(v,bool) else None
def t(v):
    if v is None: return None
    if isinstance(v,(int,float)):
        return str(int(v)) if float(v).is_integer() else str(v)
    s=str(v).strip(); return s or None
def sqlnum(v): return 'NULL' if v is None else repr(round(v,4))
def sqlstr(v): return 'NULL' if v is None else "'"+str(v).replace("'","''")+"'"
def sqldate(v):
    return "'"+v.date().isoformat()+"'" if isinstance(v,datetime.datetime) else 'NULL'

rows=[]
lot=None
for r in range(1, ws.max_row+1):
    b=ws.cell(r,2).value
    if b=='Lot No':
        c=ws.cell(r,3).value
        lot=t(c)
        continue
    if b in ('PRODUCTION REPORT','DPR No.'):
        continue
    product=clean_product(ws.cell(r,4).value)
    amount=n(ws.cell(r,19).value)
    buyer_c, buyer_raw=clean_buyer(ws.cell(r,20).value)
    # keep genuine sellable lines: real USD amount, or a buyer+grade pairing
    actual_grade = t(ws.cell(r,7).value)
    if not product or lot is None: continue
    is_sale = (amount is not None and amount > 0) or (buyer_c is not None and actual_grade is not None)
    if not is_sale: continue
    rec=dict(
        lot_no=lot, dpr_no=t(ws.cell(r,2).value), dpr_date=ws.cell(r,3).value,
        product=product, rm_count=t(ws.cell(r,5).value), rm_qty=n(ws.cell(r,6).value),
        actual_grade=t(ws.cell(r,7).value), target_grade=t(ws.cell(r,8).value),
        frozen_count=t(ws.cell(r,9).value), packing=t(ws.cell(r,10).value),
        actual_glaze=n(ws.cell(r,11).value), target_glaze=n(ws.cell(r,12).value),
        net_per_frozen=t(ws.cell(r,13).value), cases=n(ws.cell(r,14).value),
        total_qty=n(ws.cell(r,16).value), base_rate=n(ws.cell(r,17).value),
        calibrated_rate=n(ws.cell(r,18).value), amount_usd=amount,
        buyer=buyer_c, buyer_raw=buyer_raw,
        frozen_yield=n(ws.cell(r,21).value), net_yield=n(ws.cell(r,22).value),
        rm_weight=n(ws.cell(r,23).value), source_row=r,
    )
    rows.append(rec)

cols=['lot_no','dpr_no','dpr_date','product','rm_count','rm_qty','actual_grade','target_grade',
 'frozen_count','packing','actual_glaze','target_glaze','net_per_frozen','cases','total_qty',
 'base_rate','calibrated_rate','amount_usd','buyer','buyer_raw','frozen_yield','net_yield','rm_weight','source_row']
def cell(rec,k):
    v=rec[k]
    if k=='dpr_date': return sqldate(v)
    if k=='source_row': return str(v)
    if isinstance(v,float) or v is None and k in ('rm_qty','actual_glaze','target_glaze','cases','total_qty','base_rate','calibrated_rate','amount_usd','frozen_yield','net_yield','rm_weight'):
        return sqlnum(v)
    if k in ('rm_qty','actual_glaze','target_glaze','cases','total_qty','base_rate','calibrated_rate','amount_usd','frozen_yield','net_yield','rm_weight'):
        return sqlnum(v)
    return sqlstr(v)

vals=[]
for rec in rows:
    vals.append('('+','.join(cell(rec,k) for k in cols)+')')

with open('/Users/fahdnaveed/Documents/Protech Audit Files/Protech PWA/_lots.sql','w') as f:
    f.write('delete from public.lot_economics;\n')
    f.write('insert into public.lot_economics ('+','.join(cols)+') values\n')
    f.write(',\n'.join(vals)+';\n')

# summary
buyers=sorted(set(r['buyer'] for r in rows if r['buyer']))
lots=sorted(set(r['lot_no'] for r in rows if r['lot_no']))
tot_usd=sum(r['amount_usd'] or 0 for r in rows)
print('lines:',len(rows),'lots:',len(lots),'buyers:',len(buyers))
print('total USD:', round(tot_usd))
print('buyers:', buyers)
