import openpyxl, datetime, json, urllib.request

SRC = '/Users/fahdnaveed/Downloads/LOT Analysis.xlsx'
URL = 'https://ymoewukjsrbggchqzfbc.supabase.co/rest/v1/lot_economics'
KEY = 'sb_publishable_ziyCxVbFtRqjUaGmBocFOA_FkmxYmIz'

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['LOT Analysis']

BUYER = {
    'MY FOOD':'My Food','MY FOODS':'My Food',
    'VIMA':'Vima','CORPORACION ALIMENTARIA VIMA S.L.':'Vima',
    'SEA BLUE, UK':'Sea Blue, UK','SEA BLUE,UK':'Sea Blue, UK','SEABLUE, UK':'Sea Blue, UK',
    'SEA PRIDE':'Sea Pride','AL HAMOOR':'Al Hamoor','AH-CHOU':'Ah-Chou',
    'QUIRCH':'Quirch','QUICH':'Quirch','GENERAL':'General','GLOBAL':'Global',
    'OCEANIS':'Oceanis','LUCKY FISH':'Lucky Fish','LUCKY':'Lucky Fish',
    'VASSILIOU':'Vassiliou','SCHROEDER':'Schroeder','THALASA':'Thalassa',
    'THALASSA SEAFOODS NV/SA':'Thalassa','PESCANOVA':'Pescanova','FEINFROST':'Feinfrost',
    'NORDIC':'Nordic','IST':'IST','SEAFOOD':'Seafood','AZHARI':'Azhari',
    'EURL':'EURL','BML':'BML',
}
def clean_buyer(v):
    if v is None: return None, None
    raw = str(v).strip()
    if not raw: return None, None
    key = ' '.join(raw.upper().split())
    return BUYER.get(key, raw.title()), raw
def clean_product(v):
    if v is None: return None
    return ' '.join(str(v).strip().upper().split()) or None
def n(v):
    return round(float(v),4) if isinstance(v,(int,float)) and not isinstance(v,bool) else None
def t(v):
    if v is None: return None
    if isinstance(v,(int,float)):
        return str(int(v)) if float(v).is_integer() else str(v)
    s=str(v).strip(); return s or None
def d(v):
    return v.date().isoformat() if isinstance(v,datetime.datetime) else None

rows=[]; lot=None
for r in range(1, ws.max_row+1):
    b=ws.cell(r,2).value
    if b=='Lot No':
        lot=t(ws.cell(r,3).value); continue
    if b in ('PRODUCTION REPORT','DPR No.'): continue
    product=clean_product(ws.cell(r,4).value)
    amount=n(ws.cell(r,19).value)
    buyer_c, buyer_raw=clean_buyer(ws.cell(r,20).value)
    actual_grade=t(ws.cell(r,7).value)
    if not product or lot is None: continue
    is_sale=(amount is not None and amount>0) or (buyer_c is not None and actual_grade is not None)
    if not is_sale: continue
    rows.append(dict(
        lot_no=lot, dpr_no=t(ws.cell(r,2).value), dpr_date=d(ws.cell(r,3).value),
        product=product, rm_count=t(ws.cell(r,5).value), rm_qty=n(ws.cell(r,6).value),
        actual_grade=actual_grade, target_grade=t(ws.cell(r,8).value),
        frozen_count=t(ws.cell(r,9).value), packing=t(ws.cell(r,10).value),
        actual_glaze=n(ws.cell(r,11).value), target_glaze=n(ws.cell(r,12).value),
        net_per_frozen=t(ws.cell(r,13).value), cases=n(ws.cell(r,14).value),
        total_qty=n(ws.cell(r,16).value), base_rate=n(ws.cell(r,17).value),
        calibrated_rate=n(ws.cell(r,18).value), amount_usd=amount,
        buyer=buyer_c, buyer_raw=buyer_raw,
        frozen_yield=n(ws.cell(r,21).value), net_yield=n(ws.cell(r,22).value),
        rm_weight=n(ws.cell(r,23).value), source_row=r,
    ))

def req(method, body=None, prefer=None):
    h={'apikey':KEY,'Authorization':'Bearer '+KEY,'Content-Type':'application/json'}
    if prefer: h['Prefer']=prefer
    data=json.dumps(body).encode() if body is not None else None
    rq=urllib.request.Request(URL if method!='DELETE' else URL+'?source_row=gte.0', data=data, headers=h, method=method)
    with urllib.request.urlopen(rq) as resp:
        return resp.status, resp.read().decode()

# wipe then insert in batches of 200
print('delete:', req('DELETE', prefer='return=minimal')[0])
for i in range(0, len(rows), 200):
    batch=rows[i:i+200]
    st,_=req('POST', batch, prefer='return=minimal')
    print(f'  batch {i//200+1}: {len(batch)} rows -> {st}')
print('posted', len(rows), 'rows;', len(set(x['lot_no'] for x in rows)), 'lots;',
      len(set(x['buyer'] for x in rows if x['buyer'])), 'buyers; USD',
      round(sum(x['amount_usd'] or 0 for x in rows)))
