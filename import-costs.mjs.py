import openpyxl, datetime, sys

SRC = '/Users/fahdnaveed/Downloads/DAILY COST ANALYSIS - MAY  2026.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb['APRIL']

def num(r, c):
    v = ws.cell(r, c).value
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None

def sql(v):
    return 'NULL' if v is None else repr(round(v, 4)) if isinstance(v, float) else str(v)

cols = {
    'power_normal':2,'power_peak':3,'power_offpeak':4,'power_md':5,'power_total':6,'power_per_kg':7,
    'water_etp_qty':8,'water_etp_cost':9,'water_tanker_litres':10,'water_tanker_cost':11,
    'water_japan':12,'water_japan_cost':13,'water_total':14,'water_per_kg':15,
    'firewood_qty':16,'firewood_amt':17,'firewood_per_kg':18,
    'diesel_qty':19,'diesel_amt':20,'diesel_per_kg':21,
    'peeling_cost':22,'peeling_per_kg':23,
    'wages_total':24,'wages_per_kg':25,'salary_total':26,'salary_per_kg':27,
    'total_expense':28,
    'prod_plate':29,'prod_iqf':31,'prod_blast':33,'prod_aqua':35,'prod_dolphin':36,'prod_ghan':37,'prod_total':38,
}

rows = []
for r in range(6, 37):  # May 1..31
    d = ws.cell(r, 1).value
    if not isinstance(d, datetime.datetime):
        continue
    rec = {k: num(r, c) for k, c in cols.items()}
    # recompute cost_per_kg (sheet had #REF!/#DIV0 errors)
    cpk = None
    if rec['total_expense'] and rec['prod_total'] and rec['prod_total'] > 0:
        cpk = rec['total_expense'] / rec['prod_total']
    keys = list(cols.keys()) + ['cost_date', 'cost_per_kg']
    vals = [sql(rec[k]) for k in cols] + ["'%s'" % d.date().isoformat(), sql(cpk)]
    rows.append('(' + ','.join(vals) + ')')

collist = ','.join(list(cols.keys()) + ['cost_date', 'cost_per_kg'])
out = []
out.append('delete from public.daily_costs;')
out.append('insert into public.daily_costs (%s) values\n%s;' % (collist, ',\n'.join(rows)))

# process charges (names only; rate blank -> NULL)
charges = []
for r in range(49, 63):
    name = ws.cell(r, 49).value  # AW
    if name and str(name).strip() and str(name).strip().lower() != 'proposed charges':
        charges.append((str(name).strip(), r - 48))
out.append('delete from public.process_charges;')
cvals = ",\n".join("('%s', NULL, 'per kg', %d)" % (n.replace("'", "''"), so) for n, so in charges)
out.append("insert into public.process_charges (process_name, rate, unit, sort_order) values\n%s;" % cvals)

with open('/Users/fahdnaveed/Documents/Protech Audit Files/Protech PWA/_costs.sql', 'w') as f:
    f.write('\n\n'.join(out))
print('days:', len(rows), 'charges:', len(charges))
